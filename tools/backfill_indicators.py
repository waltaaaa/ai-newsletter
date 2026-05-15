"""
backfill_indicators.py — Backfill 5 years of indicator data to indicator_history table.

One-time/occasional utility. Run: python tools/backfill_indicators.py

Sources (all free APIs, $0 cost):
  - Bank of Canada Valet API: overnight rate, GoC yields, prime rate
  - Statistics Canada WDS: CPI, unemployment, employment, GDP (national + provincial)
  - Yahoo Finance: commodities, indices, FX, crypto, Canadian mining/agriculture
  - FRED (St. Louis Fed): US/UK macro, IMF commodity prices, bond spreads
  - ECB Statistical Data Warehouse: EU rates, CPI, unemployment
  - Ontario Economic Accounts (OEA): quarterly GDP components
  - Quebec ISQ: monthly/quarterly economic indicators
"""

import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import os
import json
import time
from datetime import datetime, timedelta, date

# Add project root to path so we can import db.py
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from dotenv import load_dotenv
load_dotenv()

import requests
from db import init_db, save_indicator


def backfill_boc(conn, years=5):
    """Backfill Bank of Canada series: overnight rate, GoC yields."""
    print("\n[BOC] Backfilling Bank of Canada indicators...")

    # IDs must match INDICATOR_CATALOG in app.js for frontend queries
    series = {
        'overnight_rate': 'V39079',
        'goc_2y_yield': 'BD.CDN.2YR.DQ.YLD',
        'goc_3y_yield': 'BD.CDN.3YR.DQ.YLD',
        'goc_5y_yield': 'BD.CDN.5YR.DQ.YLD',
        'goc_7y_yield': 'BD.CDN.7YR.DQ.YLD',
        'goc_10y_yield': 'BD.CDN.10YR.DQ.YLD',
        'goc_long_yield': 'BD.CDN.LONG.DQ.YLD',
        'prime_rate': 'V80691335',
    }

    start_date = (date.today() - timedelta(days=years * 365)).isoformat()
    total = 0

    for indicator_name, series_id in series.items():
        try:
            url = f"https://www.bankofcanada.ca/valet/observations/{series_id}/json?start_date={start_date}"
            resp = requests.get(url, timeout=30)
            if resp.status_code != 200:
                print(f"  [BOC] {indicator_name}: HTTP {resp.status_code}")
                continue

            data = resp.json()
            observations = data.get('observations', [])
            print(f"  [BOC] {indicator_name}: {len(observations)} observations")

            count = 0
            for obs in observations:
                obs_date = obs.get('d', '')
                val = obs.get(series_id, {})
                if isinstance(val, dict):
                    val = val.get('v', '')
                if not obs_date or not val or val == '':
                    continue

                save_indicator(conn, {
                    'indicator': indicator_name,
                    'province': 'national',
                    'date': obs_date,
                    'value': str(val),
                    'unit': '%',
                    'source': 'Bank of Canada Valet',
                    'frequency': 'daily' if 'yield' in indicator_name else 'scheduled',
                    'backfilled': True,
                })
                count += 1

            total += count
            time.sleep(1)

        except Exception as e:
            print(f"  [BOC] {indicator_name} error: {e}")

    print(f"  [BOC] Total: {total} observations written")
    return total


def backfill_statcan(conn, years=5):
    """Backfill Statistics Canada: CPI and unemployment via WDS API."""
    print("\n[STATCAN] Backfilling StatCan indicators...")

    # IDs must match INDICATOR_CATALOG in app.js: 'cpi' + province='national', 'unemployment' + province
    vectors = {
        'cpi_national': '41690973',
        'unemployment_national': '2062815',
        # Provincial unemployment
        'unemployment_NL': '2063004',
        'unemployment_PE': '2063193',
        'unemployment_NS': '2063382',
        'unemployment_NB': '2063571',
        'unemployment_QC': '2063760',
        'unemployment_ON': '2063949',
        'unemployment_MB': '2064138',
        'unemployment_SK': '2064327',
        'unemployment_AB': '2064516',
        'unemployment_BC': '2064705',
        # Provincial CPI
        'cpi_NL': '41690914',
        'cpi_PE': '41690915',
        'cpi_NS': '41690916',
        'cpi_NB': '41690917',
        'cpi_QC': '41690918',
        'cpi_ON': '41690919',
        'cpi_MB': '41690920',
        'cpi_SK': '41690921',
        'cpi_AB': '41690922',
        'cpi_BC': '41690923',
        # National employment & participation rates — must match phases/data_collection.py
        # _EMPRATE_VECTOR / _PARTRATE_VECTOR. Verified 2026-04-18: v2062817 returns
        # rate (~60.6%), v2062816 returns participation rate (~64.9%). Previous
        # values (2062811/2062807) returned counts or failed (terminated).
        'employment_rate_national': '2062817',
        'participation_rate_national': '2062816',
        # Provincial rates — use unemployment vector + offset (part=+1, emp=+2)
        # to match the runtime _PROV_PARTRATE_VIDS / _PROV_EMPRATE_VIDS tables.
        'employment_rate_NL': '2063006',
        'employment_rate_PE': '2063195',
        'employment_rate_NS': '2063384',
        'employment_rate_NB': '2063573',
        'employment_rate_QC': '2063762',
        'employment_rate_ON': '2063951',
        'employment_rate_MB': '2064140',
        'employment_rate_SK': '2064329',
        'employment_rate_AB': '2064518',
        'employment_rate_BC': '2064707',
        # Provincial participation rate (unemployment vector + 1)
        'participation_rate_NL': '2063005',
        'participation_rate_PE': '2063194',
        'participation_rate_NS': '2063383',
        'participation_rate_NB': '2063572',
        'participation_rate_QC': '2063761',
        'participation_rate_ON': '2063950',
        'participation_rate_MB': '2064139',
        'participation_rate_SK': '2064328',
        'participation_rate_AB': '2064517',
        'participation_rate_BC': '2064706',
        # Quarterly real GDP
        'realGdp_national': '62305752',
        # GDP by industry (20 NAICS, monthly)
        'gdp_agriculture': '65201229',
        'gdp_mining_og': '65201236',
        'gdp_utilities': '65201254',
        'gdp_construction': '65201258',
        'gdp_manufacturing': '65201263',
        'gdp_wholesale': '65201358',
        'gdp_retail': '65201368',
        'gdp_transportation': '65201381',
        'gdp_information': '65201398',
        'gdp_finance': '65201407',
        'gdp_real_estate': '65201419',
        'gdp_professional': '65201429',
        'gdp_management': '65201441',
        'gdp_admin_waste': '65201442',
        'gdp_education': '65201452',
        'gdp_healthcare': '65201457',
        'gdp_entertainment': '65201463',
        'gdp_accommodation': '65201468',
        'gdp_other_services': '65201471',
        'gdp_public_admin': '65201476',
        # Provincial GDP (annual)
        'gdp_NL': '62464519',
        'gdp_PE': '62464824',
        'gdp_NS': '62465129',
        'gdp_NB': '62465434',
        'gdp_QC': '62465739',
        'gdp_ON': '62466044',
        'gdp_MB': '62466349',
        'gdp_SK': '62466654',
        'gdp_AB': '62466959',
        'gdp_BC': '62467264',
        # Housing-related
        'housingStarts_national': '735337',
        'building_permits': '735391',
    }

    total = 0
    wds_url = "https://www150.statcan.gc.ca/t1/wds/rest/getDataFromVectorsAndLatestNPeriods"

    for indicator_name, vector_id in vectors.items():
        try:
            payload = [{"vectorId": int(vector_id), "latestN": years * 12 + 6}]
            resp = requests.post(
                wds_url,
                json=payload,
                headers={"Content-Type": "application/json"},
                timeout=30,
            )

            if resp.status_code != 200:
                print(f"  [STATCAN] {indicator_name}: HTTP {resp.status_code}")
                continue

            data = resp.json()
            # WDS returns [{"status": "SUCCESS", "object": {"vectorDataPoint": [...]}}]
            obs_list = []
            if isinstance(data, list):
                for series in data:
                    obj = series.get('object', {})
                    for dp in obj.get('vectorDataPoint', []):
                        obs_list.append({
                            'date': dp.get('refPer', ''),
                            'value': dp.get('value', ''),
                        })
            elif isinstance(data, dict):
                for dp in data.get('object', {}).get('vectorDataPoint', []):
                    obs_list.append({
                        'date': dp.get('refPer', ''),
                        'value': dp.get('value', ''),
                    })

            # Determine province from indicator name
            PROV_CODES = ('NL', 'PE', 'NS', 'NB', 'QC', 'ON', 'MB', 'SK', 'AB', 'BC')
            parts = indicator_name.split('_')
            if len(parts) >= 2 and parts[-1] in PROV_CODES:
                province = parts[-1]
                base_indicator = '_'.join(parts[:-1])
            elif indicator_name.endswith('_national'):
                province = 'national'
                base_indicator = indicator_name.replace('_national', '')
            else:
                province = 'national'
                base_indicator = indicator_name

            print(f"  [STATCAN] {indicator_name}: {len(obs_list)} observations")

            count = 0
            # CPI vectors return raw index levels (e.g. 160.2). The runtime stores
            # cpi as YoY% by convention, so backfill must compute YoY from 13-apart
            # observations before writing — otherwise the raw index leaks into the
            # YoY slot and fails validator PROVINCE_RULES[cpi] range check.
            is_cpi = base_indicator == 'cpi'
            for i, obs in enumerate(obs_list):
                obs_date = obs['date'][:10] if obs['date'] else ''
                val = obs['value']
                if not obs_date or val == '' or val is None:
                    continue

                # Determine frequency and unit from indicator type
                freq = 'monthly'
                unit = '%'
                if 'realGdp' in indicator_name or indicator_name.startswith('gdp_') and not indicator_name.startswith('gdp_agriculture'):
                    freq = 'quarterly' if 'realGdp' in indicator_name or indicator_name.startswith('gdp_') and indicator_name.split('_')[-1] in PROV_CODES else 'monthly'
                    unit = '$M'
                if indicator_name == 'housingStarts_national' or indicator_name == 'building_permits':
                    unit = 'units'

                if is_cpi:
                    # Need 12 months of history to compute YoY
                    if i < 12:
                        continue
                    try:
                        latest = float(val)
                        year_ago = float(obs_list[i - 12]['value'])
                        if not year_ago:
                            continue
                        yoy = ((latest - year_ago) / year_ago) * 100
                        write_val = f"+{yoy:.1f}%" if yoy >= 0 else f"{yoy:.1f}%"
                    except (ValueError, TypeError, ZeroDivisionError):
                        continue
                else:
                    write_val = str(val)

                save_indicator(conn, {
                    'indicator': base_indicator,
                    'province': province,
                    'date': obs_date,
                    'value': write_val,
                    'unit': unit,
                    'source': 'Statistics Canada WDS',
                    'frequency': freq,
                    'backfilled': True,
                })
                count += 1

            total += count
            time.sleep(0.5)

        except Exception as e:
            print(f"  [STATCAN] {indicator_name} error: {e}")

    print(f"  [STATCAN] Total: {total} observations written")
    return total


def backfill_yahoo(conn, years=5):
    """Backfill Yahoo Finance commodities: oil, gold, lumber, TSX, CAD."""
    print("\n[YAHOO] Backfilling Yahoo Finance commodities...")

    try:
        import yfinance as yf
    except ImportError:
        print("  [YAHOO] yfinance not installed, skipping")
        return 0

    # IDs must match indicator_name used by pipeline + frontend
    tickers = {
        # Existing
        'wti': 'CL=F',
        'gold': 'GC=F',
        'lumber': 'LBR=F',  # 2026-04-18: LBS=F delisted, LBR=F working
        'tsx_composite': '^GSPTSE',
        'tsx': '^GSPTSE',
        'cadusd': 'CADUSD=X',
        'natural_gas': 'NG=F',
        'copper': 'HG=F',
        # Equity indices
        'sp500': '^GSPC',
        'djia': '^DJI',
        'nasdaq': '^IXIC',
        'ftse100': '^FTSE',
        'dax': '^GDAXI',
        'nikkei225': '^N225',
        # FX pairs
        'eurusd': 'EURUSD=X',
        'usdcny': 'USDCNY=X',
        'usdjpy': 'USDJPY=X',
        # Commodities
        'brent': 'BZ=F',
        'silver': 'SI=F',
        'platinum': 'PL=F',
        'palladium': 'PA=F',
        'aluminum': 'ALI=F',
        'wheat': 'ZW=F',
        'corn': 'ZC=F',
        'soybeans': 'ZS=F',
        'coffee': 'KC=F',
        'cocoa': 'CC=F',
        'sugar': 'SB=F',
        'cotton': 'CT=F',
        # Missing from original backfill — in pipeline TICKER_MAP
        'coal': 'MTF=F',
        'propane': 'PN=F',
        'rice': 'ZR=F',
        'soybean_oil': 'ZL=F',
        'soybean_meal': 'ZM=F',
        # Crypto
        'bitcoin': 'BTC-USD',
        'ethereum': 'ETH-USD',
        # Shipping (Baltic Dry Index proxy)
        'dry_bulk_shipping': 'BDRY',
        # Canadian mining/agriculture
        'potash_nutrien': 'NTR.TO',
        'cameco_uranium': 'CCO.TO',
        'sprott_uranium': 'U-UN.TO',
        'canola': 'RS=F',
    }

    total = 0
    for indicator_name, ticker in tickers.items():
        try:
            data = yf.download(ticker, period=f"{years}y", progress=False)
            if data is None or len(data) == 0:
                print(f"  [YAHOO] {indicator_name}: no data")
                continue

            # Sample weekly (every 5th trading day) to reduce document count
            sampled = data.iloc[::5]
            print(f"  [YAHOO] {indicator_name}: {len(sampled)} weekly samples from {len(data)} daily")

            count = 0
            for idx, row in sampled.iterrows():
                obs_date = idx.strftime('%Y-%m-%d') if hasattr(idx, 'strftime') else str(idx)[:10]
                try:
                    val = float(row['Close'].iloc[0]) if hasattr(row['Close'], 'iloc') else float(row['Close'])
                except (TypeError, ValueError, IndexError):
                    continue

                save_indicator(conn, {
                    'indicator': indicator_name,
                    'province': 'national',
                    'date': obs_date,
                    'value': str(round(val, 2)),
                    'unit': 'USD' if indicator_name != 'tsx_composite' else 'CAD',
                    'source': 'Yahoo Finance',
                    'frequency': 'weekly',
                    'backfilled': True,
                })
                count += 1

            total += count
            time.sleep(1)

        except Exception as e:
            print(f"  [YAHOO] {indicator_name} error: {e}")

    print(f"  [YAHOO] Total: {total} observations written")
    return total


def backfill_oea(conn, years=5):
    """Backfill Ontario Economic Accounts from data.ontario.ca XLSX."""
    print("\n[OEA] Backfilling Ontario Economic Accounts...")

    try:
        import openpyxl, io
    except ImportError:
        print("  [OEA] openpyxl not installed, skipping")
        return 0

    url = 'https://data.ontario.ca/dataset/bfda561c-060b-4737-9bdc-1d8599662df1/resource/98f4bed6-2481-4c0b-83e5-009c35bfeb5f/download/oea_data_tables_english.xlsx'
    try:
        resp = requests.get(url, timeout=60)
        if resp.status_code != 200:
            print(f"  [OEA] HTTP {resp.status_code}")
            return 0
        print(f"  [OEA] Downloaded {len(resp.content)} bytes")
    except Exception as e:
        print(f"  [OEA] Download error: {e}")
        return 0

    wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
    total = 0
    cutoff_year = date.today().year - years

    def _quarter_to_date(q_str):
        """Convert '2025Q3' or '2025 Q3' to '2025-07-01' (first day of quarter)."""
        if not q_str:
            return None
        s = str(q_str).replace(' ', '')
        for fmt in [('Q1', '-01-01'), ('Q2', '-04-01'), ('Q3', '-07-01'), ('Q4', '-10-01')]:
            if fmt[0] in s.upper():
                yr = s.upper().replace(fmt[0], '').strip()
                try:
                    if int(yr) < cutoff_year:
                        return None
                    return yr + fmt[1]
                except ValueError:
                    return None
        return None

    def _extract_oea_series(sheet_name, row_indices, indicator_names, unit):
        """Extract time series from an OEA sheet. Row indices are 1-based."""
        nonlocal total
        try:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
        except Exception as e:
            print(f"  [OEA] Sheet {sheet_name} error: {e}")
            return

        # Row 0 (first row) has date headers starting from column 1
        header_row = rows[0]
        date_cols = []
        for col_idx, cell in enumerate(header_row):
            if col_idx == 0:
                continue
            d = _quarter_to_date(cell)
            if d:
                date_cols.append((col_idx, d))

        if not date_cols:
            print(f"  [OEA] No date columns found in {sheet_name}")
            return

        for row_idx, indicator_name in zip(row_indices, indicator_names):
            if row_idx - 1 >= len(rows):
                print(f"  [OEA] {indicator_name}: row {row_idx} out of range")
                continue
            data_row = rows[row_idx - 1]
            count = 0
            for col_idx, obs_date in date_cols:
                if col_idx >= len(data_row):
                    continue
                val = data_row[col_idx]
                if val is None or val == '' or val == '-':
                    continue
                try:
                    float(val)
                except (TypeError, ValueError):
                    continue

                save_indicator(conn, {
                    'indicator': indicator_name,
                    'province': 'ON',
                    'date': obs_date,
                    'value': str(val),
                    'unit': unit,
                    'source': 'Ontario Economic Accounts',
                    'frequency': 'quarterly',
                    'backfilled': True,
                })
                count += 1

            total += count
            print(f"  [OEA] {indicator_name}: {count} observations")

    # Table 3: Real GDP expenditure (chained 2017$, quarterly)
    _extract_oea_series('Table 3', [2, 4, 16, 18], [
        'on_real_consumption', 'on_real_household', 'on_real_gov_expenditure', 'on_real_capital_investment'
    ], '$M')

    # Table 3 % changes: Row 3, 5, 17, 19
    _extract_oea_series('Table 3', [3, 5, 17, 19], [
        'on_consumption_pct', 'on_household_pct', 'on_gov_expenditure_pct', 'on_capital_investment_pct'
    ], '%')

    # Table 5: Trade nominal (quarterly)
    _extract_oea_series('Table 5', [2, 8], ['on_exports', 'on_imports'], '$M')

    # Table 5 % changes: Row 3, 9
    _extract_oea_series('Table 5', [3, 9], ['on_exports_pct', 'on_imports_pct'], '%')

    # Table 15: GDP by industry real (chained 2017$, quarterly, starts 1997)
    _extract_oea_series('Table 15', [2], ['on_gdp_goods'], '$M')
    _extract_oea_series('Table 15', [3], ['on_gdp_goods_pct'], '%')

    wb.close()
    print(f"  [OEA] Total: {total} observations written")
    return total


def backfill_isq(conn, years=5):
    """Backfill Quebec ISQ indicators from statistique.quebec.ca XLSX."""
    print("\n[ISQ] Backfilling Quebec ISQ indicators...")

    try:
        import openpyxl, io
    except ImportError:
        print("  [ISQ] openpyxl not installed, skipping")
        return 0

    url = 'https://statistique.quebec.ca/fr/fichier/principaux-indicateurs-economiques-conjoncturels-donnees-mensuelles-trimestrielles-et-annuelles-excel.xlsx'
    try:
        resp = requests.get(url, timeout=60)
        if resp.status_code != 200:
            print(f"  [ISQ] HTTP {resp.status_code}")
            return 0
        print(f"  [ISQ] Downloaded {len(resp.content)} bytes")
    except Exception as e:
        print(f"  [ISQ] Download error: {e}")
        return 0

    wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
    ws = wb['indicat']
    rows = list(ws.iter_rows(values_only=True))
    total = 0

    # --- QUARTERLY SECTION (rows 4-21 in spreadsheet = indices 3-20) ---
    year_row = rows[3]
    qtr_row = rows[4]

    # Build quarterly date map: column index -> ISO date
    qtr_dates = {}
    current_year = None
    qtr_map = {'I': '-01-01', 'II': '-04-01', 'III': '-07-01', 'IV': '-10-01'}
    cutoff_year = date.today().year - years

    for col in range(1, len(year_row)):
        yr = year_row[col]
        qtr = str(qtr_row[col]).strip() if col < len(qtr_row) and qtr_row[col] else ''

        if yr and str(yr).strip().isdigit():
            current_year = str(yr).strip()

        if current_year and qtr in qtr_map:
            if int(current_year) >= cutoff_year:
                qtr_dates[col] = current_year + qtr_map[qtr]

    # Quarterly indicators to extract: (row_index_0based, indicator_id, unit)
    # Row indices realigned 2026-05-15 to the current ISQ workbook layout: the
    # 'indicat' quarterly-levels block shifted so government final consumption
    # is row 9 and business gross fixed capital formation is row 10 (were 10/11).
    # Verified against statistique.quebec.ca (Comptes économiques trimestriels).
    quarterly_indicators = [
        (6,  'qc_real_gdp', '$M'),
        (8,  'qc_household_consumption', '$M'),
        (9,  'qc_gov_consumption', '$M'),
        (10, 'qc_business_investment', '$M'),
        (12, 'qc_exports', '$M'),
        (13, 'qc_imports', '$M'),
        (15, 'qc_nominal_gdp', '$M'),
        (16, 'qc_compensation', '$M'),
        (19, 'qc_household_income', '$M'),
    ]

    for row_idx, indicator_name, unit in quarterly_indicators:
        if row_idx >= len(rows):
            continue
        data_row = rows[row_idx]
        count = 0
        for col, obs_date in qtr_dates.items():
            if col >= len(data_row):
                continue
            val = data_row[col]
            if val is None or val == '' or val == '-':
                continue
            try:
                float(val)
            except (TypeError, ValueError):
                continue

            save_indicator(conn, {
                'indicator': indicator_name,
                'province': 'QC',
                'date': obs_date,
                'value': str(round(float(val), 2)),
                'unit': unit,
                'source': 'Institut de la statistique du Québec',
                'frequency': 'quarterly',
                'backfilled': True,
            })
            count += 1

        total += count
        print(f"  [ISQ] {indicator_name}: {count} quarterly observations")

    # Quarterly % changes (rows 52-66 = indices 51-65)
    quarterly_pct = [
        (52, 'qc_real_gdp_pct', '%'),
        (54, 'qc_household_consumption_pct', '%'),
        (55, 'qc_gov_consumption_pct', '%'),
        (56, 'qc_business_investment_pct', '%'),
        (58, 'qc_exports_pct', '%'),
        (59, 'qc_imports_pct', '%'),
    ]

    for row_idx, indicator_name, unit in quarterly_pct:
        if row_idx >= len(rows):
            continue
        data_row = rows[row_idx]
        count = 0
        for col, obs_date in qtr_dates.items():
            if col >= len(data_row):
                continue
            val = data_row[col]
            if val is None or val == '' or val == '-':
                continue
            try:
                float(val)
            except (TypeError, ValueError):
                continue

            save_indicator(conn, {
                'indicator': indicator_name,
                'province': 'QC',
                'date': obs_date,
                'value': str(val),
                'unit': unit,
                'source': 'Institut de la statistique du Québec',
                'frequency': 'quarterly',
                'backfilled': True,
            })
            count += 1

        total += count
        print(f"  [ISQ] {indicator_name}: {count} quarterly % observations")

    # --- MONTHLY SECTION (rows 23-43 = indices 22-42) ---
    month_year_row = rows[22]
    month_name_row = rows[23]

    month_map = {
        'Janvier': '-01-01', 'F\xe9vrier': '-02-01', 'Fevrier': '-02-01',
        'Mars': '-03-01', 'Avril': '-04-01', 'Mai': '-05-01', 'Juin': '-06-01',
        'Juillet': '-07-01', 'Ao\xfbt': '-08-01', 'Aout': '-08-01',
        'Septembre': '-09-01', 'Sept.': '-09-01', 'Octobre': '-10-01',
        'Novembre': '-11-01', 'D\xe9cembre': '-12-01', 'Decembre': '-12-01',
    }

    # Build monthly date map
    monthly_dates = {}
    current_year = None
    for col in range(1, len(month_year_row)):
        yr = month_year_row[col]
        if yr and str(yr).strip().isdigit():
            current_year = str(yr).strip()

        month_name = str(month_name_row[col]).strip() if col < len(month_name_row) and month_name_row[col] else ''

        if current_year and month_name in month_map:
            if int(current_year) >= cutoff_year:
                monthly_dates[col] = current_year + month_map[month_name]

    # Monthly indicators: (row_index_0based, indicator_id, unit)
    monthly_indicators = [
        (24, 'qc_monthly_gdp', '$M'),
        (25, 'qc_intl_exports', '$M'),
        (26, 'qc_intl_imports', '$M'),
        (27, 'qc_housing_starts', 'units'),
        (28, 'qc_bldg_permits_res', '$M'),
        (29, 'qc_bldg_permits_nonres', '$M'),
        (30, 'qc_manufacturing_sales', '$M'),
        (31, 'qc_wholesale_sales', '$M'),
        (32, 'qc_retail_sales', '$M'),
        (33, 'qc_weekly_earnings', '$'),
        (34, 'qc_labour_force', 'K'),
        (35, 'qc_employment', 'K'),
        (36, 'qc_employment_ft', 'K'),
        (37, 'qc_employment_pt', 'K'),
        (38, 'qc_unemployment_rate', '%'),
        (39, 'qc_participation_rate', '%'),
        (40, 'qc_employment_rate', '%'),
        (41, 'qc_cpi', 'index'),
    ]

    for row_idx, indicator_name, unit in monthly_indicators:
        if row_idx >= len(rows):
            continue
        data_row = rows[row_idx]
        count = 0
        for col, obs_date in monthly_dates.items():
            if col >= len(data_row):
                continue
            val = data_row[col]
            if val is None or val == '' or val == '-':
                continue
            try:
                float(val)
            except (TypeError, ValueError):
                continue

            save_indicator(conn, {
                'indicator': indicator_name,
                'province': 'QC',
                'date': obs_date,
                'value': str(round(float(val), 2)),
                'unit': unit,
                'source': 'Institut de la statistique du Québec',
                'frequency': 'monthly',
                'backfilled': True,
            })
            count += 1

        total += count
        print(f"  [ISQ] {indicator_name}: {count} monthly observations")

    wb.close()
    print(f"  [ISQ] Total: {total} observations written")
    return total


def backfill_fred(conn, years=5):
    """Backfill FRED: US macro, IMF commodity prices, bond spreads."""
    print("\n[FRED] Backfilling FRED indicators...")

    start = (date.today() - timedelta(days=years * 365)).isoformat()
    total = 0

    # (indicator_name, series_id, unit, frequency, province)
    series = [
        # US macro (existing in pipeline)
        ('us_fed_funds', 'FEDFUNDS', '%', 'monthly', 'US'),
        ('us_unemployment', 'UNRATE', '%', 'monthly', 'US'),
        ('us_cpi_index', 'CPIAUCSL', 'index', 'monthly', 'US'),
        ('us_real_gdp', 'GDPC1', '$B', 'quarterly', 'US'),
        # UK
        ('uk_bank_rate', 'BOEBRBA', '%', 'daily', 'UK'),
        # IMF commodity prices via FRED (monthly)
        ('iron_ore', 'PIORECRUSDM', 'USD/t', 'monthly', 'national'),
        ('nickel', 'PNICKUSDM', 'USD/t', 'monthly', 'national'),
        ('zinc', 'PZINCUSDM', 'USD/t', 'monthly', 'national'),
        ('tin', 'PTINUSDM', 'USD/t', 'monthly', 'national'),
        ('lead', 'PLEADUSDM', 'USD/t', 'monthly', 'national'),
        ('lng_asia', 'PNGASJPUSDM', 'USD/MMBtu', 'monthly', 'national'),
        # Bond/credit spreads (daily)
        ('ig_spread', 'BAMLC0A0CM', '%', 'daily', 'US'),
        ('hy_spread', 'BAMLH0A0HYM2', '%', 'daily', 'US'),
        ('yield_curve_10y2y', 'T10Y2Y', '%', 'daily', 'US'),
    ]

    for indicator_name, series_id, unit, freq, province in series:
        try:
            url = (f"https://fred.stlouisfed.org/graph/fredgraph.csv"
                   f"?id={series_id}&observation_start={start}")
            resp = requests.get(url, timeout=45, headers={
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            })
            if resp.status_code != 200:
                print(f"  [FRED] {indicator_name}: HTTP {resp.status_code}")
                continue

            count = 0
            # Sample: daily series every 5th row to reduce volume
            sample_rate = 5 if freq == 'daily' else 1
            row_idx = 0
            for line in resp.text.strip().split('\n'):
                if line.startswith('DATE') or line.startswith('observation'):
                    continue
                row_idx += 1
                if sample_rate > 1 and row_idx % sample_rate != 0:
                    continue
                parts = line.split(',')
                if len(parts) < 2:
                    continue
                obs_date = parts[0].strip()
                val = parts[1].strip()
                if not obs_date or not val or val == '.':
                    continue
                try:
                    float(val)
                except ValueError:
                    continue

                save_indicator(conn, {
                    'indicator': indicator_name,
                    'province': province,
                    'date': obs_date,
                    'value': val,
                    'unit': unit,
                    'source': 'FRED',
                    'frequency': 'weekly' if freq == 'daily' else freq,
                    'backfilled': True,
                })
                count += 1

            total += count
            print(f"  [FRED] {indicator_name}: {count} observations")
            time.sleep(1)

        except Exception as e:
            print(f"  [FRED] {indicator_name} error: {e}")

    print(f"  [FRED] Total: {total} observations written")
    return total


def backfill_ecb(conn, years=5):
    """Backfill ECB: deposit rate, HICP, EU unemployment."""
    print("\n[ECB] Backfilling ECB indicators...")

    start_period = (date.today() - timedelta(days=years * 365)).strftime('%Y-%m')
    total = 0

    # (indicator_name, dataflow, key, unit, frequency)
    series = [
        ('ecb_deposit_rate', 'FM', 'B.U2.EUR.4F.KR.DFR.LEV', '%', 'daily'),
        ('eu_hicp', 'ICP', 'M.U2.N.000000.4.ANR', '%', 'monthly'),
        ('eu_unemployment', 'STS', 'M.I8.S.UNEH.RTT000.4.000', '%', 'monthly'),
    ]

    for indicator_name, dataflow, key, unit, freq in series:
        try:
            url = (f"https://data-api.ecb.europa.eu/service/data/{dataflow}/{key}"
                   f"?format=csvdata&startPeriod={start_period}")
            resp = requests.get(url, timeout=30)
            if resp.status_code != 200:
                print(f"  [ECB] {indicator_name}: HTTP {resp.status_code}")
                continue

            import csv
            import io
            reader = csv.DictReader(io.StringIO(resp.text))
            count = 0
            sample_rate = 5 if freq == 'daily' else 1
            row_idx = 0
            for row in reader:
                row_idx += 1
                if sample_rate > 1 and row_idx % sample_rate != 0:
                    continue
                period = row.get('TIME_PERIOD', '')
                val = row.get('OBS_VALUE', '')
                if not period or not val:
                    continue
                # Normalize period to ISO date
                if len(period) == 7:  # YYYY-MM
                    period += '-01'
                elif len(period) == 4:  # YYYY
                    period += '-01-01'

                save_indicator(conn, {
                    'indicator': indicator_name,
                    'province': 'EU',
                    'date': period,
                    'value': val,
                    'unit': unit,
                    'source': 'ECB',
                    'frequency': 'weekly' if freq == 'daily' else freq,
                    'backfilled': True,
                })
                count += 1

            total += count
            print(f"  [ECB] {indicator_name}: {count} observations")
            time.sleep(1)

        except Exception as e:
            print(f"  [ECB] {indicator_name} error: {e}")

    print(f"  [ECB] Total: {total} observations written")
    return total


if __name__ == "__main__":
    print("=" * 60)
    print("INDICATOR HISTORY BACKFILL — 5 years")
    print("=" * 60)

    conn = init_db()

    boc_count = backfill_boc(conn, years=5)
    statcan_count = backfill_statcan(conn, years=5)
    yahoo_count = backfill_yahoo(conn, years=5)
    fred_count = backfill_fred(conn, years=5)
    ecb_count = backfill_ecb(conn, years=5)
    oea_count = backfill_oea(conn, years=5)
    isq_count = backfill_isq(conn, years=5)

    conn.close()

    grand_total = boc_count + statcan_count + yahoo_count + fred_count + ecb_count + oea_count + isq_count
    print(f"\n{'=' * 60}")
    print(f"BACKFILL COMPLETE")
    print(f"  BoC:     {boc_count} observations")
    print(f"  StatCan: {statcan_count} observations")
    print(f"  Yahoo:   {yahoo_count} observations")
    print(f"  FRED:    {fred_count} observations")
    print(f"  ECB:     {ecb_count} observations")
    print(f"  OEA:     {oea_count} observations")
    print(f"  ISQ:     {isq_count} observations")
    print(f"  TOTAL:   {grand_total} observations")
    print(f"{'=' * 60}")
