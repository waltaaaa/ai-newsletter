"""
backfill_projects.py -- Standalone government data ingestion for clean project rebuild.

Downloads and parses 6 government open data sources:
1. Infrastructure Canada CSV (~20K projects, federal funding)
2. NRCan Major Projects Inventory XLSX (~500 projects, energy/mining/forest)
3. Ontario Builds CSV (~5,900 projects, provincial infrastructure)
4. BC Major Projects Inventory XLSX (~1,000 projects, all sectors)
5. Alberta Major Projects API (~970 projects, all sectors)
6. Quebec Infrastructure Plan CSV (~3,100 projects, provincial infrastructure)

Normalizes all to canonical schema, applies GDP threshold filters,
cross-source deduplicates, and outputs a review CSV before loading.

Usage:
    python tools/backfill_projects.py                    # Parse, filter, output review CSV
    python tools/backfill_projects.py --load             # Load approved review CSV into DB
    python tools/backfill_projects.py --load --export    # Load + regenerate frontend JSON
"""

import argparse
import csv
import json
import logging
import os
import re
import sys
from datetime import datetime, date
from pathlib import Path

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from normalize import normalize_province, normalize_status, parse_value
from pipeline_config import SECTOR_CANONICAL_MAP, PROVINCE_GDP_THRESHOLDS, infer_naics

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

DATA_DIR = Path(__file__).parent.parent / "data" / "backfill"
REVIEW_CSV = Path(__file__).parent.parent / "backfill_review.csv"

# Date boundaries
LOOKBACK_START = date(2024, 3, 15)
COMPLETED_BEFORE = date(2025, 1, 1)

# ── Helpers ─────────────────────────────────────────────────────────────────

def _sector_lookup(raw):
    """Map a government source sector/category to canonical 18-sector key."""
    if not raw:
        return "infrastructure"  # default
    key = raw.strip().lower()
    return SECTOR_CANONICAL_MAP.get(key, None)


def _parse_date(raw):
    """Try to parse a date string in various formats. Returns YYYY-MM-DD or ''."""
    if not raw or not isinstance(raw, str):
        return ""
    raw = raw.strip()
    if not raw or raw in ("N/A", "TBD", "Unknown", "—"):
        return ""
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S.%f",
                "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d", "%B %d, %Y", "%b %d, %Y",
                "%Y"):
        try:
            return datetime.strptime(raw[:len(raw)], fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    # Try quarter format: 2025-Q2 → 2025-04-01
    qm = re.match(r"(\d{4})-Q(\d)", raw)
    if qm:
        year, q = int(qm.group(1)), int(qm.group(2))
        month = {1: 1, 2: 4, 3: 7, 4: 10}.get(q, 1)
        return f"{year}-{month:02d}-01"
    return ""


def _fmt_value(pv):
    """Format a parsed numeric value for display. Uses B for billions."""
    if not pv or pv <= 0:
        return "Not disclosed"
    if pv >= 1_000_000_000:
        return f"${pv/1_000_000_000:.1f}B"
    if pv >= 1_000_000:
        return f"${pv/1_000_000:.0f}M"
    if pv >= 1_000:
        return f"${pv/1_000:.0f}K"
    return f"${pv:,.0f}"


def _date_to_date(s):
    """Parse YYYY-MM-DD string to date object, or None."""
    if not s:
        return None
    try:
        return datetime.strptime(s[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def _build_evidence(url, source_name):
    """Build a single evidence entry dict."""
    return {
        "url": url,
        "source_type": "government_registry",
        "name": source_name,
        "date": datetime.now().strftime("%Y-%m-%d"),
        "authority": "government",
    }


def _passes_threshold(parsed_value, province, source="", category=""):
    """Check if a project meets the province GDP threshold.

    For major project inventories (NRCan, BC MPI, Alberta, Quebec): no-value projects
    pass as unconfirmed, since these sources only track significant projects.

    For funding program lists (Infrastructure Canada): a confirmed value above
    threshold is required, since these include everything down to $50K.

    For Ontario Builds: confirmed value required EXCEPT for Transit and Health care
    categories, which are curated major infrastructure even without disclosed budgets.
    """
    # Ontario Builds Transit and Health care — allow unconfirmed (inherently major)
    ONTARIO_MAJOR_CATEGORIES = {"transit", "health care"}

    if parsed_value is None:
        # Infrastructure Canada always requires confirmed value
        if "Infrastructure Canada" in source:
            return False
        # Ontario Builds requires confirmed value EXCEPT for major categories
        if "Ontario Builds" in source:
            return category.lower().strip() in ONTARIO_MAJOR_CATEGORIES
        return True  # major project inventory → include as unconfirmed
    threshold = PROVINCE_GDP_THRESHOLDS.get(province, 500_000_000)
    return parsed_value >= threshold


def _passes_date_filter(announcement_date, completion_date, status):
    """Check date-based filters:
    - Must have activity within 2-year lookback window
    - Exclude completed before 2025-01-01
    """
    # If completed and we have a completion date, check it
    if status == "Complete" and completion_date:
        comp = _date_to_date(completion_date)
        if comp and comp < COMPLETED_BEFORE:
            return False

    # If we have an announcement/start date, check lookback
    if announcement_date:
        ann = _date_to_date(announcement_date)
        if ann and ann < LOOKBACK_START:
            # Old project — only include if not complete
            if status == "Complete":
                return False
    return True


# ── Source Parsers ──────────────────────────────────────────────────────────

def parse_infrastructure_canada():
    """Parse Infrastructure Canada CSV — federal infrastructure funding."""
    filepath = DATA_DIR / "infra_canada.csv"
    if not filepath.exists():
        logger.warning("Infrastructure Canada CSV not found")
        return []

    projects = []
    with open(filepath, encoding="latin1") as f:
        reader = csv.DictReader(f)
        for row in reader:
            title = (row.get("Title") or "").strip()
            if not title:
                continue

            raw_prov = row.get("Province/Territory", "")
            primary, additional = normalize_province(raw_prov)
            if not primary or primary == "CA":
                continue

            category = row.get("Category", "")
            sector = _sector_lookup(category)
            if not sector:
                sector = "infrastructure"

            # Value: use Total Eligible Costs
            raw_value = row.get("Total Eligible Costs", "")
            pv = None
            if raw_value:
                try:
                    pv = float(raw_value.replace(",", ""))
                except ValueError:
                    pv = None
            value_text = _fmt_value(pv)

            # Dates
            approved = _parse_date(row.get("Approved Date", ""))
            start = _parse_date(row.get("Construction Start Date") or row.get("Forecasted Construction Start Date", ""))
            end = _parse_date(row.get("Construction End Date") or row.get("Forecasted Construction End Date", ""))

            # Status inference from dates
            status = "Proposed"
            if row.get("Construction Start Date"):
                status = "Under Construction"
            elif row.get("Construction End Date"):
                status = "Complete"
            elif approved:
                status = "Approved"

            location = row.get("Location", "")
            recipient = row.get("Ultimate Recipient", "")
            proj_num = row.get("Project Number", "")
            source_url = f"https://www.infrastructure.gc.ca/gmap-gcarte/index-eng.html"

            projects.append({
                "name": title,
                "province": primary,
                "provinces_additional": additional,
                "cma": location,
                "sector": sector,
                "value": value_text,
                "parsed_value": pv,
                "status": status,
                "proponent": recipient,
                "description": f"{category} project in {location}. Recipient: {recipient}." if recipient else f"{category} project in {location}.",
                "announcement_date": approved,
                "start_date": start,
                "completionDate": end,
                "evidence": [_build_evidence(source_url, "Infrastructure Canada")],
                "discovery_source": "infra_canada_backfill",
                "confidence": 0.7,
                "has_government_source": 1,
                "source": "Infrastructure Canada",
                "_category": category,
            })

    logger.info(f"Infrastructure Canada: {len(projects)} projects parsed")
    return projects


def parse_nrcan():
    """Parse NRCan Major Projects Inventory XLSX."""
    filepath = DATA_DIR / "nrcan_mpi.xlsx"
    if not filepath.exists():
        logger.warning("NRCan XLSX not found")
        return []

    projects = []
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[1]  # Project Name
        if not name:
            continue

        raw_prov = str(row[4] or "")  # P/T
        primary, additional = normalize_province(raw_prov)
        if not primary:
            continue

        sector = _sector_lookup(row[24])  # Sector
        if not sector:
            sector = "power_energy"

        # Cost in $M
        cost_m = row[22]  # Cost 2024 (CAD) {$M}
        pv = float(cost_m) * 1_000_000 if cost_m else None
        value_text = _fmt_value(pv)

        status = normalize_status(str(row[19] or "Proposed"))  # Status 2024
        proponent = str(row[2] or "")  # Company/Proponent
        location = str(row[5] or "")  # Location

        # Build description from components
        phases = str(row[3] or "")  # Phases/Components
        desc_parts = []
        if phases:
            desc_parts.append(f"Components: {phases}.")
        energy_type = row[28] or row[29]  # ENERGY-type or Cleantech-Type
        if energy_type:
            desc_parts.append(f"Type: {energy_type}.")
        if row[27]:  # Clean Technology
            desc_parts.append("Clean technology project.")
        mining_group = row[25]
        if mining_group:
            desc_parts.append(f"Mining group: {mining_group}.")
        description = " ".join(desc_parts) if desc_parts else f"NRCan {row[24] or 'energy'} project in {location}."

        source_url = "https://natural-resources.canada.ca/energy/energy-sources-distribution/canadian-energy-resource-development/major-projects-inventory/18702"

        projects.append({
            "name": str(name),
            "province": primary,
            "provinces_additional": additional,
            "cma": location,
            "sector": sector,
            "value": value_text,
            "parsed_value": pv,
            "status": status,
            "proponent": proponent,
            "description": description,
            "announcement_date": "",
            "start_date": "",
            "completionDate": "",
            "evidence": [_build_evidence(source_url, "NRCan Major Projects Inventory")],
            "discovery_source": "nrcan_backfill",
            "confidence": 0.75,
            "has_government_source": 1,
            "source": "NRCan",
        })

    wb.close()
    logger.info(f"NRCan: {len(projects)} projects parsed")
    return projects


def parse_ontario():
    """Parse Ontario Builds CSV."""
    filepath = DATA_DIR / "ontario_builds.csv"
    if not filepath.exists():
        logger.warning("Ontario Builds CSV not found")
        return []

    projects = []
    with open(filepath, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            name = (row.get("Project") or "").strip()
            if not name:
                continue

            category = row.get("Category", "")
            sector = _sector_lookup(category)
            if not sector:
                sector = "infrastructure"

            # Value
            raw_budget = row.get("Estimated Total Budget ($)", "")
            pv = None
            if raw_budget:
                try:
                    pv = float(raw_budget.replace(",", "").replace("$", ""))
                    if pv == 0:
                        pv = None
                except ValueError:
                    pv = None
            value_text = _fmt_value(pv)

            status = normalize_status(row.get("Status", "Proposed"))
            community = row.get("Community", "")
            description = (row.get("Description") or "").strip()
            target_completion = row.get("Target Completion Date", "")
            completion = _parse_date(target_completion)

            source_url = row.get("Website", "").strip()
            if not source_url:
                source_url = "https://www.ontario.ca/page/ontario-builds"

            projects.append({
                "name": name,
                "province": "ON",
                "provinces_additional": "",
                "cma": community,
                "sector": sector,
                "value": value_text,
                "parsed_value": pv,
                "status": status,
                "proponent": row.get("Supporting Ministry", ""),
                "description": description[:500] if description else f"Ontario {category} project in {community}.",
                "announcement_date": "",
                "start_date": "",
                "completionDate": completion,
                "evidence": [_build_evidence(source_url, "Ontario Builds")],
                "discovery_source": "ontario_builds_backfill",
                "confidence": 0.7,
                "has_government_source": 1,
                "source": "Ontario Builds",
                "_category": category,
            })

    logger.info(f"Ontario Builds: {len(projects)} projects parsed")
    return projects


def parse_bc():
    """Parse BC Major Projects Inventory XLSX."""
    filepath = DATA_DIR / "bc_mpi.xlsx"
    if not filepath.exists():
        logger.warning("BC MPI XLSX not found")
        return []

    projects = []
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb["mpi_dataset_q2_2025"]

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[1]  # PROJECT_NAME
        if not name:
            continue

        # Sector: try PROJECT_TYPE first (more specific), fallback to PROJECT_CATEGORY_NAME
        project_type = str(row[8] or "").strip()
        category = str(row[15] or "").strip()
        sector = _sector_lookup(project_type) or _sector_lookup(category) or "infrastructure"

        # Cost in $M
        cost_m = row[3]  # ESTIMATED_COST
        pv = float(cost_m) * 1_000_000 if cost_m else None
        value_text = _fmt_value(pv)

        status = normalize_status(str(row[13] or "Proposed"))  # PROJECT_STATUS
        developer = str(row[11] or "")  # DEVELOPER
        municipality = str(row[10] or "")  # MUNICIPALITY
        description = str(row[2] or "")[:500]  # PROJECT_DESCRIPTION

        start = _parse_date(str(row[29] or ""))  # STANDARDIZED_START_DATE
        end = _parse_date(str(row[30] or ""))  # STANDARDIZED_COMPLETION_DATE
        first_entry = _parse_date(str(row[37] or ""))  # FIRST_ENTRY_DATE

        website = str(row[36] or "").strip()
        source_url = website if website and website.startswith("http") else "https://www2.gov.bc.ca/gov/content/employment-business/economic-development/industry/bc-major-projects-inventory"

        projects.append({
            "name": str(name),
            "province": "BC",
            "provinces_additional": "",
            "cma": municipality,
            "sector": sector,
            "value": value_text,
            "parsed_value": pv,
            "status": status,
            "proponent": developer,
            "description": description if description else f"BC {category} project in {municipality}.",
            "announcement_date": first_entry,
            "start_date": start,
            "completionDate": end,
            "evidence": [_build_evidence(source_url, "BC Major Projects Inventory")],
            "discovery_source": "bc_mpi_backfill",
            "confidence": 0.7,
            "has_government_source": 1,
            "source": "BC MPI",
        })

    wb.close()
    logger.info(f"BC MPI: {len(projects)} projects parsed")
    return projects


def parse_alberta():
    """Parse Alberta Major Projects API JSON (GeoJSON features)."""
    filepath = DATA_DIR / "alberta_projects.json"
    if not filepath.exists():
        logger.warning("Alberta JSON not found")
        return []

    with open(filepath) as f:
        data = json.load(f)

    projects = []
    for feature in data:
        props = feature.get("properties", {})
        name = props.get("name", "")
        if not name:
            continue

        sector = _sector_lookup(props.get("sector", ""))
        if not sector:
            sector = "infrastructure"

        # Cost in $M
        cost_m = props.get("cost")
        pv = float(cost_m) * 1_000_000 if cost_m else None
        value_text = _fmt_value(pv)

        # Alberta uses both stage and status
        stage = props.get("stage", "")
        status = normalize_status(stage or props.get("status", "Proposed"))

        developer = props.get("developer", "") or ""
        municipalities = props.get("municipalities", [])
        municipality = municipalities[0] if municipalities else ""

        schedule_start = _parse_date(str(props.get("schedule", "")))
        schedule_end = _parse_date(str(props.get("scheduleEnd", "")))

        friendly = props.get("friendlyName", "")
        website = props.get("website", "")
        source_url = website if website and website.startswith("http") else f"https://majorprojects.alberta.ca/details/{friendly}"

        project_type = props.get("type", "")
        description = f"{project_type}." if project_type else f"Alberta {props.get('sector', '')} project."

        projects.append({
            "name": name,
            "province": "AB",
            "provinces_additional": "",
            "cma": municipality,
            "sector": sector,
            "value": value_text,
            "parsed_value": pv,
            "status": status,
            "proponent": developer,
            "description": description,
            "announcement_date": schedule_start,
            "start_date": "",
            "completionDate": schedule_end,
            "evidence": [_build_evidence(source_url, "Alberta Major Projects")],
            "discovery_source": "alberta_backfill",
            "confidence": 0.7,
            "has_government_source": 1,
            "source": "Alberta",
        })

    logger.info(f"Alberta: {len(projects)} projects parsed")
    return projects


def parse_quebec():
    """Parse Quebec Infrastructure Plan CSV."""
    filepath = DATA_DIR / "quebec_infra.csv"
    if not filepath.exists():
        logger.warning("Quebec CSV not found")
        return []

    projects = []
    # Quebec CSV uses mixed encoding
    for enc in ("utf-8-sig", "utf-8", "latin1", "cp1252"):
        try:
            with open(filepath, encoding=enc) as f:
                reader = csv.DictReader(f)
                rows = list(reader)
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
    else:
        logger.error("Could not decode Quebec CSV")
        return []

    for row in rows:
        name = (row.get("nom_projet") or "").strip()
        if not name:
            continue

        raw_sector = row.get("secteur_activite", "")
        sector = _sector_lookup(raw_sector)
        if not sector:
            sector = "infrastructure"

        # Value: cout_total
        raw_cost = row.get("cout_total", "")
        pv = parse_value(raw_cost) if raw_cost else None
        # Quebec costs may be plain numbers in dollars
        if pv is None and raw_cost:
            try:
                cleaned = raw_cost.replace(",", "").replace(" ", "").replace("\xa0", "")
                pv = float(cleaned) if cleaned else None
            except ValueError:
                pv = None
        value_text = f"${pv/1_000_000:.0f}M" if pv and pv >= 1_000_000 else (
            f"${pv:,.0f}" if pv else "Not disclosed"
        )

        # Status from etat_avancement
        raw_etat = row.get("etat_avancement", "")
        status_map_qc = {
            "en planification": "Approved",
            "en réalisation": "Under Construction",
            "en realisation": "Under Construction",
            "en service": "Complete",
            "à l'étude": "Under Review",
            "a l'etude": "Under Review",
        }
        status = status_map_qc.get(raw_etat.lower().strip(), "Proposed")

        description = (row.get("description") or "")[:500]
        location = row.get("localisation", "")
        region = row.get("region", "")
        # Clean region: "06 – Montréal" → "Montréal"
        if region and "–" in region:
            region = region.split("–", 1)[1].strip()
        elif region and "-" in region:
            region = region.split("-", 1)[1].strip()

        completion = _parse_date(row.get("date_fin_mise_en_service", ""))
        proponent = row.get("organisme", "")
        proj_num = row.get("no_projet", "")

        source_url = f"https://www.quebec.ca/gouvernement/politiques-orientations/plan-quebecois-infrastructures"

        projects.append({
            "name": name,
            "province": "QC",
            "provinces_additional": "",
            "cma": location or region,
            "sector": sector,
            "value": value_text,
            "parsed_value": pv,
            "status": status,
            "proponent": proponent or "",
            "description": description if description else f"Quebec {raw_sector} project in {location or region}.",
            "announcement_date": "",
            "start_date": "",
            "completionDate": completion,
            "evidence": [_build_evidence(source_url, "Quebec Infrastructure Plan")],
            "discovery_source": "quebec_backfill",
            "confidence": 0.7,
            "has_government_source": 1,
            "source": "Quebec",
        })

    logger.info(f"Quebec: {len(projects)} projects parsed")
    return projects


# ── Filtering ───────────────────────────────────────────────────────────────

def apply_filters(projects):
    """Apply GDP threshold, date lookback, and data quality filters."""
    filtered = []
    stats = {"threshold": 0, "date": 0, "no_name": 0, "passed": 0}

    for p in projects:
        name = p.get("name", "").strip()
        if not name or len(name) < 5:
            stats["no_name"] += 1
            continue

        province = p.get("province", "")
        pv = p.get("parsed_value")
        status = p.get("status", "Proposed")

        # GDP threshold filter
        source = p.get("source", "")
        category = p.get("_category", "")
        if not _passes_threshold(pv, province, source, category):
            stats["threshold"] += 1
            continue

        # Date filter
        ann = p.get("announcement_date", "")
        comp = p.get("completionDate", "")
        if not _passes_date_filter(ann, comp, status):
            stats["date"] += 1
            continue

        stats["passed"] += 1
        filtered.append(p)

    logger.info(
        f"Filtering: {len(projects)} -> {stats['passed']} passed "
        f"({stats['threshold']} below threshold, {stats['date']} date-filtered, "
        f"{stats['no_name']} bad name)"
    )
    return filtered


# ── Cross-source Dedup ──────────────────────────────────────────────────────

def _dedup_key(name, province):
    """Generate a simple dedup key from name + province."""
    n = re.sub(r"[^a-z0-9]", "", name.lower())
    return f"{n}__{province}"


def cross_source_dedup(projects):
    """Deduplicate projects appearing in multiple government sources.
    Merges evidence, takes highest status progression."""
    from normalize import CANONICAL_STATUSES

    STATUS_ORDER = {s: i for i, s in enumerate(CANONICAL_STATUSES)}

    seen = {}
    for p in projects:
        key = _dedup_key(p["name"], p["province"])
        if key in seen:
            existing = seen[key]
            # Merge evidence
            existing_urls = {e.get("url") for e in existing["evidence"]}
            for e in p["evidence"]:
                if e.get("url") not in existing_urls:
                    existing["evidence"].append(e)
            # Take higher status
            ex_order = STATUS_ORDER.get(existing["status"], 0)
            new_order = STATUS_ORDER.get(p["status"], 0)
            if new_order > ex_order:
                existing["status"] = p["status"]
            # Take value if existing has none
            if not existing.get("parsed_value") and p.get("parsed_value"):
                existing["parsed_value"] = p["parsed_value"]
                existing["value"] = p["value"]
            # Take description if longer
            if len(p.get("description", "")) > len(existing.get("description", "")):
                existing["description"] = p["description"]
            # Take dates if missing
            for field in ("announcement_date", "start_date", "completionDate"):
                if not existing.get(field) and p.get(field):
                    existing[field] = p[field]
            # Merge sources
            existing["source"] = f"{existing['source']}, {p['source']}"
            # Boost confidence for multi-source
            existing["confidence"] = min(1.0, existing["confidence"] + 0.1)
        else:
            seen[key] = p

    deduped = list(seen.values())
    logger.info(f"Dedup: {len(projects)} -> {len(deduped)} unique projects")
    return deduped


# ── Review CSV ──────────────────────────────────────────────────────────────

REVIEW_FIELDS = [
    "name", "province", "provinces_additional", "sector", "value", "parsed_value",
    "status", "proponent", "cma", "description",
    "announcement_date", "start_date", "completionDate",
    "confidence", "source", "evidence_count",
]


def write_review_csv(projects, filepath=REVIEW_CSV):
    """Write projects to review CSV for user approval."""
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=REVIEW_FIELDS, extrasaction="ignore")
        writer.writeheader()
        for p in sorted(projects, key=lambda x: (x["province"], -(x.get("parsed_value") or 0))):
            p["evidence_count"] = len(p.get("evidence", []))
            writer.writerow(p)
    logger.info(f"Review CSV written: {filepath} ({len(projects)} projects)")


# ── Load into DB ────────────────────────────────────────────────────────────

def load_into_db(projects):
    """Load projects into the database via db.upsert_project()."""
    import db

    conn = db.init_db()
    loaded = 0
    skipped = 0

    for p in projects:
        try:
            key = db.upsert_project(conn, p)
            if key:
                loaded += 1
            else:
                skipped += 1
        except Exception as e:
            logger.warning(f"Failed to upsert {p['name']}: {e}")
            skipped += 1

    conn.close()
    logger.info(f"Loaded: {loaded} projects, skipped: {skipped}")
    return loaded


def run_export():
    """Run the dashboard export to regenerate frontend JSON."""
    from tools.export_dashboard import export_all
    import db

    conn = db.init_db()
    export_all(conn)
    conn.close()
    logger.info("Frontend JSON export complete")


# ── Main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Backfill projects from government open data")
    parser.add_argument("--load", action="store_true", help="Load review CSV into database")
    parser.add_argument("--export", action="store_true", help="Regenerate frontend JSON after load")
    parser.add_argument("--source", help="Parse only one source (infra_canada, nrcan, ontario, bc, alberta, quebec)")
    parser.add_argument("--refresh", action="store_true",
                        help="Monthly refresh: re-download sources, parse, diff against DB, upsert new/changed projects")
    args = parser.parse_args()

    if args.refresh:
        # Monthly refresh: re-download, parse, filter, dedup, load directly (no review CSV)
        import subprocess
        print("Downloading fresh government data...")
        dl_cmds = [
            ("infra_canada.csv", "https://www.infrastructure.gc.ca/alt-format/opendata/project-list-with-forcast-dates-liste-de-projets-avec-dates-prevu-en.csv"),
            ("nrcan_mpi.xlsx", "https://ftp.maps.canada.ca/pub/nrcan_rncan/Natural-resources_Ressources-naturelles/major_projects_inventory/MPI_2024_Active_Projects_en.xlsx"),
            ("alberta_projects.json", "https://majorprojects.alberta.ca/api/MajorProjects"),
            ("bc_mpi.xlsx", "https://www2.gov.bc.ca/assets/gov/employment-business-and-economic-development/economic-development/develop-economic-sectors/mpi/mpi-2025/mpi_dataset_q2_2025.xlsx"),
            ("quebec_infra.csv", "https://www.donneesquebec.ca/recherche/dataset/67d85a7a-10af-4af0-b4da-abc64cfd735a/resource/35de39af-84c5-4478-9857-2b4c65564c15/download/donnees_ouvertes_tb_20251219.csv"),
            # NOTE: Ontario Builds URL changes with datestamps — update manually or skip
        ]
        os.makedirs(DATA_DIR, exist_ok=True)
        for fname, url in dl_cmds:
            try:
                subprocess.run(["curl", "-sL", "-o", str(DATA_DIR / fname), url],
                              timeout=60, check=True)
                print(f"  Downloaded {fname}")
            except Exception as e:
                print(f"  Failed to download {fname}: {e}")

        # Parse all sources
        all_projects = []
        for name, fn in {"infra_canada": parse_infrastructure_canada, "nrcan": parse_nrcan,
                         "bc": parse_bc, "alberta": parse_alberta, "quebec": parse_quebec}.items():
            try:
                all_projects.extend(fn())
            except Exception as e:
                logger.error(f"Error parsing {name}: {e}")

        # Filter and dedup
        filtered = apply_filters(all_projects)
        deduped = cross_source_dedup(filtered)

        # Load directly into DB (no review step for monthly refresh)
        loaded = load_into_db(deduped)
        print(f"\nMonthly refresh: {loaded} projects upserted")

        if args.export:
            run_export()
        return

    if args.load:
        # Load from review CSV
        if not REVIEW_CSV.exists():
            logger.error(f"Review CSV not found: {REVIEW_CSV}")
            return
        projects = []
        with open(REVIEW_CSV, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                # Reconstruct evidence from source name
                source = row.get("source", "Government")
                row["evidence"] = [_build_evidence(
                    "https://www.infrastructure.gc.ca" if "Infrastructure Canada" in source
                    else "https://natural-resources.canada.ca" if "NRCan" in source
                    else "https://www.ontario.ca/page/ontario-builds" if "Ontario" in source
                    else "https://www2.gov.bc.ca" if "BC" in source
                    else "https://majorprojects.alberta.ca" if "Alberta" in source
                    else "https://www.quebec.ca",
                    source
                )]
                row["has_government_source"] = 1
                row["discovery_source"] = "government_backfill"
                # Parse numeric fields
                try:
                    row["parsed_value"] = float(row["parsed_value"]) if row.get("parsed_value") else None
                except (ValueError, TypeError):
                    row["parsed_value"] = None
                try:
                    row["confidence"] = float(row.get("confidence", 0.7))
                except (ValueError, TypeError):
                    row["confidence"] = 0.7
                projects.append(row)

        loaded = load_into_db(projects)
        print(f"\nLoaded {loaded} projects into database.")

        if args.export:
            run_export()
        return

    # Parse all sources
    parsers = {
        "infra_canada": parse_infrastructure_canada,
        "nrcan": parse_nrcan,
        "ontario": parse_ontario,
        "bc": parse_bc,
        "alberta": parse_alberta,
        "quebec": parse_quebec,
    }

    all_projects = []
    if args.source:
        if args.source not in parsers:
            logger.error(f"Unknown source: {args.source}. Available: {list(parsers.keys())}")
            return
        all_projects = parsers[args.source]()
    else:
        for name, fn in parsers.items():
            try:
                all_projects.extend(fn())
            except Exception as e:
                logger.error(f"Error parsing {name}: {e}")

    # Report raw counts by source
    source_counts = {}
    for p in all_projects:
        src = p.get("source", "unknown")
        source_counts[src] = source_counts.get(src, 0) + 1
    print(f"\n{'='*60}")
    print(f"RAW PARSED: {len(all_projects)} projects")
    for src, cnt in sorted(source_counts.items()):
        print(f"  {src}: {cnt}")

    # Apply filters
    filtered = apply_filters(all_projects)

    # Cross-source dedup
    deduped = cross_source_dedup(filtered)

    # Report by province
    prov_counts = {}
    for p in deduped:
        prov = p.get("province", "??")
        prov_counts[prov] = prov_counts.get(prov, 0) + 1
    print(f"\n{'='*60}")
    print(f"FINAL: {len(deduped)} projects after filtering + dedup")
    for prov in sorted(prov_counts.keys()):
        print(f"  {prov}: {prov_counts[prov]}")

    # Value distribution
    with_value = sum(1 for p in deduped if p.get("parsed_value"))
    total_value = sum(p.get("parsed_value", 0) or 0 for p in deduped)
    print(f"\n  With dollar value: {with_value}/{len(deduped)}")
    print(f"  Total value: ${total_value/1_000_000_000:.1f}B")

    # Write review CSV
    write_review_csv(deduped)
    print(f"\nReview CSV: {REVIEW_CSV}")
    print("Inspect the CSV, then run: python tools/backfill_projects.py --load --export")


if __name__ == "__main__":
    main()
