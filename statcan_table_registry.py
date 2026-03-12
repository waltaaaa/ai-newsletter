"""
StatCan Table Registry Builder
==============================
Pulls every table from the Statistics Canada WDS getAllCubesListLite endpoint
and produces a clean XLSX registry with: table name, table ID, link, frequency,
coverage (national/provincial/CMA/etc.), and focus area.

Single GET call — no API key required. Runs in ~10-20 seconds.

Usage:
    python statcan_table_registry.py                    # full registry
    python statcan_table_registry.py --active-only      # exclude archived tables
    python statcan_table_registry.py --csv              # also export CSV
    python statcan_table_registry.py --focus labour     # filter to one focus area

Output: statcan_table_registry.xlsx (+ .csv if --csv flag)

Requirements: requests, openpyxl, pandas
    pip install requests openpyxl pandas
"""

import argparse
import json
import re
import sys
from datetime import datetime

import pandas as pd
import requests

# ---------------------------------------------------------------------------
# WDS endpoints
# ---------------------------------------------------------------------------
BASE = "https://www150.statcan.gc.ca/t1/wds/rest"
URL_CUBES_LITE = f"{BASE}/getAllCubesListLite"
URL_CODESETS = f"{BASE}/getCodeSets"

# ---------------------------------------------------------------------------
# Frequency code mapping (from getCodeSets frequencyCode)
# Hardcoded fallback — the script also tries to pull live from getCodeSets
# ---------------------------------------------------------------------------
FREQ_FALLBACK = {
    1: "Daily",
    2: "Weekly",
    3: "Every 2 weeks",
    4: "Twice monthly",
    5: "3 times/month",
    6: "Monthly",
    7: "Every 2 months",
    8: "Quarterly",
    9: "3 times/year",
    10: "Semi-annual",
    11: "3 times in 2 years",
    12: "Annual",
    13: "Every 2 years",
    14: "Every 3 years",
    15: "Every 4 years",
    16: "Every 5 years",
    17: "Every 10 years",
    18: "Occasional",
}

# ---------------------------------------------------------------------------
# Subject code → Focus area mapping (first 2 digits of subjectCode)
# Source: StatCan subject taxonomy
# ---------------------------------------------------------------------------
SUBJECT_TO_FOCUS = {
    "10": "Banking and financial statistics",
    "11": "Agriculture",
    "12": "Fishing",
    "13": "Manufacturing",
    "14": "Construction",
    "15": "Wholesale and retail trade",
    "16": "Energy",
    "17": "Transportation",
    "18": "International trade",
    "19": "Service industries",
    "20": "Labour",
    "21": "Income and wealth",
    "22": "Government finance",
    "23": "Prices and price indexes",
    "24": "Public finance",
    "25": "Business performance and ownership",
    "26": "Justice and crime",
    "27": "Immigration and ethnocultural diversity",
    "28": "Education, training and learning",
    "29": "Information and communications technology",
    "31": "Government",
    "32": "Business and consumer surveys",
    "33": "Travel and tourism",
    "34": "Infrastructure",
    "35": "Courts and corrections",
    "36": "National accounts and GDP",
    "37": "Health",
    "38": "Culture and recreation",
    "39": "Children, youth and families",
    "40": "Society and community",
    "41": "Population and demographics",
    "42": "Languages",
    "43": "Ethnic diversity and immigration",
    "44": "Aboriginal peoples",
    "45": "Environment",
    "46": "Science and technology",
    "47": "Digital economy",
    "48": "Housing",
    "98": "Census",
    "99": "Reference and classification",
}


def format_pid(product_id: int) -> str:
    """Format raw productId (e.g. 14100287) → '14-10-0287-01'."""
    s = str(product_id).zfill(8)
    return f"{s[0:2]}-{s[2:4]}-{s[4:8]}-{s[8:]}" if len(s) > 8 else f"{s[0:2]}-{s[2:4]}-{s[4:8]}"


def build_link(product_id: int) -> str:
    """Build the StatCan website link for a table."""
    pid = str(product_id)
    return f"https://www150.statcan.gc.ca/t1/tbl1/en/tv.action?pid={pid}"


def infer_coverage(title: str) -> str:
    """Infer geographic coverage from the table title using keyword patterns."""
    t = title.lower()
    coverage_signals = [
        (r"\bcensus metropolitan area", "CMA"),
        (r"\bcma\b", "CMA"),
        (r"\bcensus agglomeration", "Census agglomeration"),
        (r"\beconomic region", "Economic region"),
        (r"\bhealth region", "Health region"),
        (r"\bcensus division", "Census division"),
        (r"\bcensus tract", "Census tract"),
        (r"\bcensus subdivision", "Census subdivision"),
        (r"\bfederal electoral district", "Federal electoral district"),
        (r"\bforward sortation area", "Forward sortation area"),
        (r"\bmunicipal", "Municipal"),
        (r"\bcit(?:y|ies)\b", "Municipal"),
        (r"\bprovince\b|\bprovinc(?:ial|es)\b|\bterritor(?:y|ies|ial)\b", "Provincial/territorial"),
        (r"\bby province\b", "Provincial/territorial"),
        (r"\bnational\b|\bcanada\b", "National"),
        (r"\binternational\b|\bcountry\b|\bcountries\b|\bworld\b", "International"),
    ]
    matches = []
    for pattern, label in coverage_signals:
        if re.search(pattern, t):
            matches.append(label)
    if not matches:
        return "National (default)"
    # Return the most specific match (first hit = most specific since list is ordered)
    return matches[0]


def resolve_focus(subject_codes: list) -> str:
    """Map subjectCode list to focus area labels. Returns the primary (first) match."""
    if not subject_codes:
        return "Unclassified"
    focuses = []
    for code in subject_codes:
        prefix = str(code)[:2]
        label = SUBJECT_TO_FOCUS.get(prefix)
        if label and label not in focuses:
            focuses.append(label)
    return "; ".join(focuses) if focuses else "Unclassified"


def fetch_frequency_map() -> dict:
    """Try to pull live frequency code descriptions from getCodeSets."""
    try:
        resp = requests.get(URL_CODESETS, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        freq_map = {}
        for item in data:
            if item.get("codeSetTitleEn", "").lower().startswith("frequency"):
                for code in item.get("codeSets", []):
                    freq_map[code["codeId"]] = code["codeDescEn"]
        if freq_map:
            return freq_map
    except Exception:
        pass
    return FREQ_FALLBACK


def fetch_all_tables() -> list:
    """Fetch the full cube list from getAllCubesListLite."""
    print("Fetching table catalog from StatCan WDS... ", end="", flush=True)
    resp = requests.get(URL_CUBES_LITE, timeout=120)
    resp.raise_for_status()
    tables = resp.json()
    print(f"{len(tables):,} tables retrieved.")
    return tables


def build_registry(tables: list, freq_map: dict, active_only: bool = False) -> pd.DataFrame:
    """Transform raw JSON into clean registry DataFrame."""
    rows = []
    for t in tables:
        # Skip archived if requested
        archived = str(t.get("archived", "2"))
        if active_only and archived != "2":  # "2" = CURRENT
            continue

        pid = t["productId"]
        title = t.get("cubeTitleEn", "")
        freq_code = t.get("frequencyCode")
        subject_codes = t.get("subjectCode", [])

        rows.append({
            "Table Name": title,
            "Table ID": format_pid(pid),
            "Product ID (raw)": pid,
            "CANSIM ID": t.get("cansimId", ""),
            "Link": build_link(pid),
            "Frequency": freq_map.get(freq_code, f"Code {freq_code}"),
            "Coverage": infer_coverage(title),
            "Focus": resolve_focus(subject_codes),
            "Subject Codes": "; ".join(str(c) for c in (subject_codes or [])),
            "Survey Codes": "; ".join(str(c) for c in (t.get("surveyCode") or [])),
            "Start Date": t.get("cubeStartDate", ""),
            "End Date": t.get("cubeEndDate", ""),
            "Last Release": t.get("releaseTime", ""),
            "Status": "Current" if archived == "2" else ("Archived" if archived == "1" else f"Code {archived}"),
        })

    df = pd.DataFrame(rows)
    df.sort_values(["Focus", "Table Name"], inplace=True, ignore_index=True)
    return df


def write_xlsx(df: pd.DataFrame, path: str):
    """Write registry to a formatted XLSX file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "StatCan Table Registry"

    # Styles
    hdr_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2E4057")
    data_font = Font(name="Arial", size=9)
    link_font = Font(name="Arial", size=9, color="1155CC", underline="single")
    border = Border(
        bottom=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
    )
    wrap = Alignment(wrap_text=True, vertical="top")

    col_widths = {
        "Table Name": 60,
        "Table ID": 16,
        "Product ID (raw)": 14,
        "CANSIM ID": 12,
        "Link": 50,
        "Frequency": 14,
        "Coverage": 24,
        "Focus": 36,
        "Subject Codes": 14,
        "Survey Codes": 12,
        "Start Date": 12,
        "End Date": 12,
        "Last Release": 20,
        "Status": 10,
    }

    # Headers
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 14)

    ws.row_dimensions[1].height = 28

    # Data
    alt_fill = PatternFill("solid", fgColor="F5F7FA")
    for row_idx, row in df.iterrows():
        excel_row = row_idx + 2
        fill = alt_fill if excel_row % 2 == 0 else None
        for col_idx, (col_name, val) in enumerate(row.items(), 1):
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.font = link_font if col_name == "Link" else data_font
            cell.alignment = wrap
            cell.border = border
            if fill:
                cell.fill = fill

    ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
    ws.freeze_panes = "A2"

    # --- Summary sheet ---
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "StatCan Table Registry — Summary"
    ws2["A1"].font = Font(name="Arial", bold=True, size=14)
    ws2["A3"] = "Generated:"
    ws2["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws2["A4"] = "Total tables:"
    ws2["B4"] = len(df)
    ws2["A5"] = "Current (active):"
    ws2["B5"] = len(df[df["Status"] == "Current"])
    ws2["A6"] = "Archived:"
    ws2["B6"] = len(df[df["Status"] == "Archived"])
    for r in range(3, 7):
        ws2.cell(row=r, column=1).font = Font(name="Arial", size=11)
        ws2.cell(row=r, column=2).font = Font(name="Arial", size=11, bold=True)

    ws2["A8"] = "Tables by Focus Area"
    ws2["A8"].font = Font(name="Arial", bold=True, size=11)
    focus_counts = df["Focus"].value_counts().sort_index()
    for i, (focus, count) in enumerate(focus_counts.items(), 9):
        ws2.cell(row=i, column=1, value=focus).font = Font(name="Arial", size=10)
        ws2.cell(row=i, column=2, value=count).font = Font(name="Arial", size=10)

    freq_start = i + 2
    ws2.cell(row=freq_start, column=1, value="Tables by Frequency").font = Font(name="Arial", bold=True, size=11)
    freq_counts = df["Frequency"].value_counts().sort_index()
    for j, (freq, count) in enumerate(freq_counts.items(), freq_start + 1):
        ws2.cell(row=j, column=1, value=freq).font = Font(name="Arial", size=10)
        ws2.cell(row=j, column=2, value=count).font = Font(name="Arial", size=10)

    cov_start = j + 2
    ws2.cell(row=cov_start, column=1, value="Tables by Coverage Level").font = Font(name="Arial", bold=True, size=11)
    cov_counts = df["Coverage"].value_counts().sort_index()
    for k, (cov, count) in enumerate(cov_counts.items(), cov_start + 1):
        ws2.cell(row=k, column=1, value=cov).font = Font(name="Arial", size=10)
        ws2.cell(row=k, column=2, value=count).font = Font(name="Arial", size=10)

    ws2.column_dimensions["A"].width = 45
    ws2.column_dimensions["B"].width = 12

    wb.save(path)
    print(f"XLSX saved: {path}")


def write_json(df: pd.DataFrame, path: str):
    """Write compact JSON for the frontend Data Explorer.

    Format per entry: {t: tableId, n: name, k: keywords, c: category, f: freq, g: geo}
    where keywords = focus area text for search matching.
    """
    # Map full frequency names to short codes
    freq_short = {
        "Daily": "D", "Weekly": "W", "Monthly": "M", "Quarterly": "Q",
        "Annual": "A", "Semi-annual": "S", "Every 2 months": "E",
        "Occasional": "O",
    }
    records = []
    for _, row in df.iterrows():
        table_id = row["Table ID"]
        freq = freq_short.get(row["Frequency"], row["Frequency"][:1] if row["Frequency"] else "")
        records.append({
            "t": table_id,
            "n": row["Table Name"],
            "k": row["Focus"],
            "c": row["Focus"].split(";")[0].strip() if row["Focus"] else "",
            "f": freq,
            "g": row["Coverage"],
        })
    with open(path, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, separators=(",", ":"))
    print(f"JSON saved: {path} ({len(records):,} entries)")


def main():
    parser = argparse.ArgumentParser(description="Build a StatCan table registry from WDS.")
    parser.add_argument("--active-only", action="store_true", help="Exclude archived tables")
    parser.add_argument("--csv", action="store_true", help="Also export CSV")
    parser.add_argument("--json", action="store_true", help="Export compact JSON for frontend")
    parser.add_argument("--json-path", type=str, default=None, help="Custom JSON output path(s), comma-separated")
    parser.add_argument("--focus", type=str, default=None, help="Filter to tables matching this focus keyword (case-insensitive)")
    parser.add_argument("--output", type=str, default="statcan_table_registry", help="Output filename stem (no extension)")
    args = parser.parse_args()

    freq_map = fetch_frequency_map()
    tables = fetch_all_tables()
    df = build_registry(tables, freq_map, active_only=args.active_only)

    if args.focus:
        mask = df["Focus"].str.contains(args.focus, case=False, na=False)
        df = df[mask].reset_index(drop=True)
        print(f"Filtered to {len(df):,} tables matching focus '{args.focus}'")

    xlsx_path = f"{args.output}.xlsx"
    write_xlsx(df, xlsx_path)

    if args.csv:
        csv_path = f"{args.output}.csv"
        df.to_csv(csv_path, index=False)
        print(f"CSV saved: {csv_path}")

    if args.json or args.json_path:
        if args.json_path:
            for p in args.json_path.split(","):
                write_json(df, p.strip())
        else:
            write_json(df, f"{args.output}.json")

    print(f"\nDone. {len(df):,} tables in registry.")
    print(f"\nFocus area breakdown:")
    for focus, count in df["Focus"].value_counts().sort_index().items():
        print(f"  {focus}: {count}")


if __name__ == "__main__":
    main()
